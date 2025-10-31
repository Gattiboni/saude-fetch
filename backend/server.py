import json
import os
import uuid
import logging
import io
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorDatabase
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

LIVE_DEBUG_ENABLED = os.getenv("LIVE_DEBUG", "false").lower() == "true"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("saude_fetch")
if LIVE_DEBUG_ENABLED:
    logger.setLevel(logging.DEBUG)

from drivers.driver_manager import manager as driver_manager
from drivers.base import BaseDriver, DriverResult
from utils.logger import JobLogger
from utils.auth import create_access_token, verify_token, check_credentials, AuthError
from utils.validators import validate_cpf_cnpj
from db.cache import Cache
from bson import ObjectId

AMIL_BROWSER_ENGINE = os.getenv("AMIL_ENGINE", "firefox").lower()
CNPJ_PIPELINE_ENABLED = False
_manual_pages: Dict[str, dict] = {}



# --- APP PRINCIPAL ---
app = FastAPI(title="saude-fetch API", version="0.2.0")

# --- CORS ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- NOVO ENDPOINT: recarregar mappings ---
@app.post("/api/mappings/reload")
def reload_mappings():
    try:
        driver_manager.reload()
        return {"detail": "Mappings recarregados com sucesso"}
    except Exception as e:
        return {"detail": f"Erro ao recarregar mappings: {e}"}


# --- Diretórios ---
BASE_DIR = os.path.dirname(__file__)
UPLOAD_DIR = os.path.join(BASE_DIR, "data", "uploads")
EXPORT_DIR = os.path.join(BASE_DIR, "data", "exports")
LOGS_DIR = os.path.join(BASE_DIR, "data", "logs")
ERRORS_DIR = os.path.join(BASE_DIR, "data", "errors")
LAST_RUN_LOG = os.path.join(LOGS_DIR, "last_run.log")

if CNPJ_PIPELINE_ENABLED:
    CNPJ_EXPORT_PATH: Optional[str] = os.path.join(EXPORT_DIR, "sulamerica_cnpj.xlsx")
    CNPJ_LOG_FILE: Optional[str] = os.path.join(LOGS_DIR, "sulamerica_cnpj.log")
    CNPJ_ERROR_DIR = os.path.join(ERRORS_DIR, "sulamerica")
else:
    CNPJ_EXPORT_PATH = None
    CNPJ_LOG_FILE = None
    CNPJ_ERROR_DIR = os.path.join(ERRORS_DIR, "sulamerica")

_required_dirs = [UPLOAD_DIR, EXPORT_DIR, LOGS_DIR, ERRORS_DIR]
if CNPJ_PIPELINE_ENABLED:
    _required_dirs.append(CNPJ_ERROR_DIR)

for d in _required_dirs:
    os.makedirs(d, exist_ok=True)

# --- Mongo ---
mongo_client: Optional[AsyncIOMotorClient] = None
mongo_db: Optional[AsyncIOMotorDatabase] = None
MONGO_URL = os.environ.get("MONGO_URL")
MONGO_DB_NAME = os.environ.get("MONGO_DB_NAME", "saude_fetch")


# --- MODELOS ---
class JobOut(BaseModel):
    id: str
    filename: str
    type: str
    status: str
    total: int
    success: int
    error: int
    created_at: str
    completed_at: Optional[str] = None
    processed: Optional[int] = 0


class JobList(BaseModel):
    items: List[JobOut]


class CnpjRequest(BaseModel):
    cnpjs: List[str]


class AmilRunRequest(BaseModel):
    token: str
    identifiers: List[str]


async def get_db():
    global mongo_client, mongo_db
    if mongo_db is None:
        if not MONGO_URL:
            raise RuntimeError("MONGO_URL not set in environment.")
        client = AsyncIOMotorClient(MONGO_URL)
        await client.admin.command("ping")
        mongo_client = client
        mongo_db = client[MONGO_DB_NAME]
    return mongo_db


async def _find_job_doc(db: AsyncIOMotorDatabase, job_id: str) -> Optional[Dict[str, Any]]:
    doc = await db.jobs.find_one({"_id": job_id})
    if doc:
        return doc
    try:
        oid = ObjectId(job_id)
    except Exception:
        return None
    return await db.jobs.find_one({"_id": oid})


async def _find_job_doc(db: AsyncIOMotorDatabase, job_id: str) -> Optional[Dict[str, Any]]:
    doc = await db.jobs.find_one({"_id": job_id})
    if doc:
        return doc
    try:
        oid = ObjectId(job_id)
    except Exception:
        return None
    return await db.jobs.find_one({"_id": oid})


@app.on_event("shutdown")
async def shutdown_event():
    global mongo_client
    if mongo_client:
        mongo_client.close()


# --- AUTH ---
class LoginIn(BaseModel):
    username: str
    password: str


class LoginOut(BaseModel):
    token: str
    expires_in_hours: int


async def require_auth(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="missing token")
    token = authorization.split(" ", 1)[1]
    try:
        user = verify_token(token)
        return user
    except AuthError as e:
        raise HTTPException(status_code=401, detail=str(e))


@app.post("/api/auth/login", response_model=LoginOut)
async def login(data: LoginIn):
    if not check_credentials(data.username, data.password):
        raise HTTPException(status_code=401, detail="invalid credentials")
    token = create_access_token(data.username, 24)
    return LoginOut(token=token, expires_in_hours=24)


# --- HEALTH ---
@app.get("/api/health")
async def health():
    return {"status": "ok", "time": datetime.utcnow().isoformat()}


@app.get("/ping")
async def ping():
    return {"status": "ok"}


# --- JOBS ---
@app.get("/api/jobs", response_model=JobList)
async def list_jobs(user: str = Depends(require_auth)):
    db = await get_db()
    cursor = db.jobs.find({}, sort=[("created_at", -1)], limit=50)
    items = []
    async for doc in cursor:
        items.append(JobOut(
            id=str(doc.get("_id")),
            filename=doc.get("filename", ""),
            type=str(doc.get("type", "auto")),
            status=doc.get("status", "pending"),
            total=int(doc.get("total", 0)),
            success=int(doc.get("success", 0)),
            error=int(doc.get("error", 0)),
            created_at=str(doc.get("created_at", "")),
            completed_at=str(doc.get("completed_at", "")) if doc.get("completed_at") else None,
            processed=int(doc.get("processed", 0)),
        ))
    return JobList(items=items)


@app.get("/api/jobs/{job_id}", response_model=JobOut)
async def get_job(job_id: str, user: str = Depends(require_auth)):
    db = await get_db()
    doc = await _find_job_doc(db, job_id)
    doc = await _find_job_doc(db, job_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Job not found")
    return JobOut(
        id=str(doc.get("_id")),
        filename=doc.get("filename", ""),
        type=str(doc.get("type", "auto")),
        status=doc.get("status", "pending"),
        total=int(doc.get("total", 0)),
        success=int(doc.get("success", 0)),
        error=int(doc.get("error", 0)),
        created_at=str(doc.get("created_at", "")),
        completed_at=str(doc.get("completed_at", "")) if doc.get("completed_at") else None,
        processed=int(doc.get("processed", 0)),
    )


@app.get("/api/jobs/{job_id}/log")
async def download_job_log(job_id: str, user: str = Depends(require_auth)):
    db = await get_db()
    doc = await _find_job_doc(db, job_id)
    path = None
    if doc:
        path = doc.get("job_log_path")
    if not path:
        path = os.path.join(LOGS_DIR, f"{job_id}.log")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Log not found")
    return FileResponse(path, media_type="text/plain", filename=f"{job_id}.log")


@app.get("/api/jobs/{job_id}/results")
async def download_job_results(job_id: str, format: str = "json", user: str = Depends(require_auth)):
    db = await get_db()
    doc = await _find_job_doc(db, job_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Job not found")

    if format.lower() == "xlsx":
        xlsx_path = doc.get("xlsx_path") or os.path.join(EXPORT_DIR, f"{job_id}.xlsx")
        if not os.path.exists(xlsx_path):
            raise HTTPException(status_code=404, detail="XLSX not found")
        return FileResponse(
            xlsx_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{job_id}.xlsx",
        )

    cursor = db.job_results.find({"job_id": job_id})
    items = []
    async for row in cursor:
        row.pop("_id", None)
        items.append(row)

    if not items:
        raise HTTPException(status_code=404, detail="Results not found")

    return {"items": items}


@app.get("/api/jobs/{job_id}/log")
async def download_job_log(job_id: str, user: str = Depends(require_auth)):
    db = await get_db()
    doc = await _find_job_doc(db, job_id)
    path = None
    if doc:
        path = doc.get("job_log_path")
    if not path:
        path = os.path.join(LOGS_DIR, f"{job_id}.log")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Log not found")
    return FileResponse(path, media_type="text/plain", filename=f"{job_id}.log")


@app.get("/api/jobs/{job_id}/results")
async def download_job_results(job_id: str, format: str = "json", user: str = Depends(require_auth)):
    db = await get_db()
    doc = await _find_job_doc(db, job_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Job not found")

    if format.lower() == "xlsx":
        xlsx_path = doc.get("xlsx_path") or os.path.join(EXPORT_DIR, f"{job_id}.xlsx")
        if not os.path.exists(xlsx_path):
            raise HTTPException(status_code=404, detail="XLSX not found")
        return FileResponse(
            xlsx_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{job_id}.xlsx",
        )

    cursor = db.job_results.find({"job_id": job_id})
    items = []
    async for row in cursor:
        row.pop("_id", None)
        items.append(row)

    if not items:
        raise HTTPException(status_code=404, detail="Results not found")

    return {"items": items}


@app.post("/api/jobs", response_model=JobOut)
async def create_job(background_tasks: BackgroundTasks, file: UploadFile = File(...), user: str = Depends(require_auth)):
    try:
        filename = file.filename or "upload"
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".csv", ".xlsx", ".xls"]:
            raise HTTPException(status_code=400, detail="Unsupported file type. Upload CSV or Excel.")

        db = await get_db()
        job_id = str(uuid.uuid4())
        created_at = datetime.utcnow().isoformat()

        doc = {
            "_id": job_id,
            "filename": filename,
            "type": "cpf",
            "status": "pending",
            "total": 0,
            "success": 0,
            "error": 0,
            "processed": 0,
            "created_at": created_at,
            "completed_at": None,
            "export_path": None,
            "xlsx_path": None,
            "file_path": None,
            "error_message": None,
        }
        await db.jobs.insert_one(doc)

        stored_name = f"{job_id}{ext}"
        stored_path = os.path.join(UPLOAD_DIR, stored_name)
        with open(stored_path, "wb") as out:
            chunk = await file.read()
            out.write(chunk)

        await db.jobs.update_one({"_id": job_id}, {"$set": {"status": "processing", "file_path": stored_path}})

        background_tasks.add_task(process_job, job_id, stored_path, "cpf")

        return JobOut(
            id=job_id,
            filename=filename,
            type="cpf",
            status="processing",
            total=0,
            success=0,
            error=0,
            created_at=created_at,
            completed_at=None,
            processed=0,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@app.post("/api/manual/amil/start")
async def start_amil_manual(user: str = Depends(require_auth)):
    """
    Inicia navegador headful para fluxo manual da Amil.
    Retorna token que o frontend usará para rodar a busca.
    """
    from playwright.async_api import async_playwright
    import secrets

    token = secrets.token_urlsafe(8)
    pw = await async_playwright().start()
    engine = getattr(pw, AMIL_BROWSER_ENGINE, None)
    if engine is None:
        await pw.stop()
        raise HTTPException(status_code=500, detail=f"Unsupported browser engine: {AMIL_BROWSER_ENGINE}")

    launch_kwargs: Dict[str, Any] = {"headless": False}
    if AMIL_BROWSER_ENGINE == "firefox":
        launch_kwargs["slow_mo"] = 150

    browser = await engine.launch(**launch_kwargs)
    context = await browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/121.0.0.0 Safari/537.36"
        ),
        viewport={"width": 1366, "height": 768},
        locale="pt-BR",
        timezone_id="America/Sao_Paulo",
    )
    page = await context.new_page()

    _manual_pages[token] = {
        "playwright": pw,
        "browser": browser,
        "context": context,
        "page": page,
        "created_at": datetime.utcnow().isoformat(),
    }
    return {
        "token": token,
        "note": "Navegador aberto. Cole o link da Amil e carregue a página.",
    }


@app.post("/api/manual/amil/run")
async def run_amil_manual(data: AmilRunRequest, user: str = Depends(require_auth)):
    """
    Anexa o driver Amil à página aberta e executa a busca.
    """
    session = _manual_pages.get(data.token)
    if not session:
        raise HTTPException(status_code=404, detail="session not found")

    page = session["page"]

    from drivers.amil import AmilDriver

    driver = AmilDriver()
    results: List[DriverResult] = []
    for ident in data.identifiers:
        try:
            result = await driver.consult(ident, id_type="cpf", page=page)
        except Exception as exc:
            result = DriverResult(
                operator=driver.operator,
                status="erro",
                message=str(exc),
                identifier=ident,
                id_type="cpf",
            )
        results.append(result)

    return {"results": [r.__dict__ for r in results]}


@app.post("/api/manual/amil/upload")
async def manual_amil_upload(
    file: UploadFile = File(...), user: str = Depends(require_auth)
):
    """
    Recebe CSV/XLSX com CPFs, retorna lista de CPFs válidos e inválidos.
    """
    try:
        filename = file.filename or "upload"
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".csv", ".xls", ".xlsx"]:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file type. Upload CSV or Excel.",
            )

        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Arquivo vazio.")

        import pandas as pd

        if ext == ".csv":
            df = pd.read_csv(io.BytesIO(content), dtype=str)
        else:
            df = pd.read_excel(io.BytesIO(content), dtype=str)

        raw_identifiers = to_rows(df, forced_type="cpf")
        cleaned = [
            clean_identifier(x) for x in raw_identifiers if str(x).strip()
        ]

        valid: List[str] = []
        invalid: List[str] = []
        for ident in cleaned:
            if validate_cpf_cnpj(ident) and len(ident) == 11:
                valid.append(ident)
            else:
                invalid.append(ident)

        return {"total": len(cleaned), "valid": valid, "invalid": invalid}
    except HTTPException:
        raise
    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# --- CNPJ PIPELINE ---
@app.post("/api/fetch/cnpj")
async def fetch_cnpj(data: CnpjRequest, user: str = Depends(require_auth)):
    if not CNPJ_PIPELINE_ENABLED:
        raise HTTPException(
            status_code=503,
            detail="Pipeline CNPJ desativado temporariamente.",
        )

    from drivers.sulamerica import SulamericaDriver

    driver = SulamericaDriver()
    rows: List[Dict[str, Any]] = []
    ativos = 0
    inativos = 0
    erros = 0

    for raw in data.cnpjs:
        digits = clean_identifier(raw)
        captured_at = datetime.utcnow().isoformat()
        if len(digits) != 14:
            status = "invalid"
            mensagem = "CNPJ inválido"
            plano = ""
            debug = {"reason": "invalid_length"}
        else:
            try:
                result = await driver.consult(digits, "cnpj")
                status = result.status
                plano = result.plan
                mensagem = result.message or result.plan or ""
                debug = result.debug
            except Exception as exc:
                status = "erro"
                plano = ""
                mensagem = str(exc)
                debug = {"exception": str(exc)}

        if status == "ativo":
            ativos += 1
        elif status == "inativo":
            inativos += 1
        elif status not in ("invalid", "indefinido"):
            erros += 1

        row = {
            "cnpj": digits,
            "status": status,
            "mensagem_portal": mensagem,
            "plano": plano,
            "captured_at": captured_at,
            "debug": debug,
        }
        rows.append(row)
        append_cnpj_log(
            {
                "event": "cnpj_result",
                "cnpj": digits,
                "status": status,
                "mensagem": mensagem,
                "plano": plano,
                "debug": debug,
            }
        )

    if CNPJ_EXPORT_PATH:
        build_cnpj_xlsx(rows, CNPJ_EXPORT_PATH)

    summary = {
        "total": len(rows),
        "ativos": ativos,
        "inativos": inativos,
        "erros": erros,
        "resultados": [
            {
                "cnpj": format_cnpj(row["cnpj"]),
                "status": row["status"],
                "mensagem_portal": row["mensagem_portal"],
                "plano": row["plano"],
                "timestamp": row["captured_at"],
            }
            for row in rows
        ],
    }

    append_cnpj_log({"event": "cnpj_summary", **summary})

    return summary


@app.get("/api/fetch/cnpj/export")
async def download_cnpj_export(user: str = Depends(require_auth)):
    if not CNPJ_PIPELINE_ENABLED or not CNPJ_EXPORT_PATH:
        raise HTTPException(
            status_code=503,
            detail="Pipeline CNPJ desativado temporariamente.",
        )
    if not os.path.exists(CNPJ_EXPORT_PATH):
        raise HTTPException(status_code=404, detail="Export not found")
    return FileResponse(
        CNPJ_EXPORT_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="sulamerica_cnpj.xlsx",
    )


# --- HELPERS ---
def clean_identifier(s: str) -> str:
    return "".join(ch for ch in str(s) if ch.isdigit())


def format_cpf(cpf_digits: str) -> str:
    if len(cpf_digits) != 11:
        return cpf_digits
    return f"{cpf_digits[0:3]}.{cpf_digits[3:6]}.{cpf_digits[6:9]}-{cpf_digits[9:11]}"


def format_cnpj(cnpj_digits: str) -> str:
    if len(cnpj_digits) != 14:
        return cnpj_digits
    return (
        f"{cnpj_digits[0:2]}.{cnpj_digits[2:5]}.{cnpj_digits[5:8]}/"
        f"{cnpj_digits[8:12]}-{cnpj_digits[12:14]}"
    )


def detect_type(identifier: str) -> str:
    if len(identifier) == 11:
        return "cpf"
    if len(identifier) == 14:
        return "cnpj"
    return "unknown"


def to_rows(df: pd.DataFrame, forced_type: str) -> List[str]:
    if df.empty:
        return []
    for col in df.columns.tolist():
        series = df[col].dropna().astype(str)
        if not series.empty:
            return [clean_identifier(x) for x in series.tolist() if str(x).strip()]
    return []


def append_cnpj_log(record: Dict[str, Any]) -> None:
    if not CNPJ_PIPELINE_ENABLED or not CNPJ_LOG_FILE:
        return

    payload = {**record, "time": datetime.utcnow().isoformat()}
    with open(CNPJ_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def build_cnpj_xlsx(rows: List[Dict[str, Any]], out_path: Optional[str]) -> None:
    if not CNPJ_PIPELINE_ENABLED or not out_path:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Consulta CNPJ"
    ws.append(["CNPJ", "STATUS", "MENSAGEM", "PLANO", "TIMESTAMP"])
    for row in rows:
        ws.append(
            [
                format_cnpj(str(row.get("cnpj", ""))),
                row.get("status", ""),
                row.get("mensagem_portal", ""),
                row.get("plano", ""),
                row.get("captured_at", ""),
            ]
        )
    for cell in ws["A"]:
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")
    wb.save(out_path)


async def process_job(job_id: str, path: str, forced_type: str = "auto"):
    db = await get_db()
    cache = Cache(db)
    cache = Cache(db)
    job_logger = JobLogger(job_id, LOGS_DIR)
    job_started_at = datetime.utcnow().isoformat()
    detailed_entries: List[Dict[str, Any]] = []
    await db.job_results.delete_many({"job_id": job_id})
    job_logger.info("job_started", job_id=job_id, file_path=path, forced_type=forced_type)
    logger.info(f"[LIVE] Status atual do job: {job_id} - started")
    job_started_at = datetime.utcnow().isoformat()
    detailed_entries: List[Dict[str, Any]] = []
    await db.job_results.delete_many({"job_id": job_id})
    job_logger.info("job_started", job_id=job_id, file_path=path, forced_type=forced_type)
    logger.info(f"[LIVE] Status atual do job: {job_id} - started")
    try:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            df = pd.read_csv(path, dtype=str)
        else:
            df = pd.read_excel(path, dtype=str)

        raw_identifiers = to_rows(df, forced_type)
        identifiers = [clean_identifier(x) for x in raw_identifiers if str(x).strip()]
        total = len(identifiers)

        await db.jobs.update_one(
            {"_id": job_id},
            {"$set": {"total": total, "processed": 0, "success": 0, "error": 0}},
        )

        results: List[Dict[str, Any]] = []
        success = 0
        error = 0
        processed = 0
        drivers = driver_manager.drivers
        job_logger.info("identifiers_loaded", total=total)

        invalid_identifiers: List[str] = []
        grouped: Dict[str, List[str]] = defaultdict(list)
        for ident in identifiers:
            if not validate_cpf_cnpj(ident):
                invalid_identifiers.append(ident)
                continue
            itype = detect_type(ident) if forced_type == "auto" else forced_type
            if itype not in ("cpf", "cnpj"):
                invalid_identifiers.append(ident)
                continue
            grouped[itype].append(ident)

        async def update_job_progress():
            await db.jobs.update_one(
                {"_id": job_id},
                {
                    "$set": {
                        "processed": processed,
                        "success": success,
                        "error": error,
                        "total": total,
                    }
                },
            )

        for ident in invalid_identifiers:
            detail_entry = {
                "input": ident,
                "type": "invalid",
                "operator": "",
                "status": "invalid",
                "plan": "",
                "message": "identificador inválido",
                "captured_at": datetime.utcnow().isoformat(),
                "debug": {"reason": "invalid_identifier"},
            }
            results.append(detail_entry)
            detailed_entries.append(detail_entry)
            error += 1
            processed += 1
            job_logger.error("identifier_invalid", identifier=ident, id_type="invalid")
            await update_job_progress()

        identifier_meta: Dict[str, Dict[str, Any]] = {}
        results_buffer: Dict[str, List[DriverResult]] = defaultdict(list)

        async def finalize_identifier(identifier: str) -> None:
            nonlocal success, error, processed
            meta = identifier_meta.pop(identifier, {"id_type": forced_type, "expected": 0})
            entries = results_buffer.pop(identifier, [])
            if not entries:
                detail_entry = {
                    "input": identifier,
                    "type": meta.get("id_type", forced_type),
                    "operator": "",
                    "status": "erro",
                    "plan": "",
                    "message": "sem resultado",
                    "captured_at": datetime.utcnow().isoformat(),
                    "debug": {"reason": "no_result"},
                }
                results.append(detail_entry)
                detailed_entries.append(detail_entry)
                error += 1
                processed += 1
                job_logger.error(
                    "identifier_without_result",
                    identifier=identifier,
                    id_type=meta.get("id_type", forced_type),
                )
                await update_job_progress()
                return

            has_success = any(
                entry.status.lower() not in {"erro", "invalid"} for entry in entries
            )
            if has_success:
                success += 1
            else:
                error += 1
            processed += 1

            for entry in entries:
                detail_entry = {
                    "input": identifier,
                    "type": meta.get("id_type", forced_type),
                    "operator": entry.operator,
                    "status": entry.status,
                    "plan": entry.plan,
                    "message": entry.message,
                    "captured_at": entry.captured_at,
                    "debug": entry.debug,
                }
                results.append(detail_entry)
                detailed_entries.append(detail_entry)

            await update_job_progress()
            job_logger.info(
                "identifier_processed",
                identifier=identifier,
                id_type=meta.get("id_type", forced_type),
                drivers=len(entries),
                success=has_success,
            )

        async def handle_progress(
            identifier: str, driver: BaseDriver, result: DriverResult, from_cache: bool
        ) -> None:
            meta = identifier_meta.get(identifier)
            if not meta:
                return
            debug_info = dict(result.debug or {})
            if from_cache:
                debug_info["cache_hit"] = True
            result.debug = debug_info
            job_logger.info(
                "driver_result",
                identifier=identifier,
                id_type=meta.get("id_type", forced_type),
                driver=driver.name,
                status=result.status,
                plan=result.plan,
                message=result.message,
                cached=from_cache,
                debug=debug_info,
            )
            results_buffer[identifier].append(result)
            if len(results_buffer[identifier]) >= meta.get("expected", 0):
                await finalize_identifier(identifier)

        for id_type, identifiers in grouped.items():
            if not identifiers:
                continue
            active_drivers = [
                drv
                for drv in drivers
                if id_type in getattr(drv, "supported_id_types", ("cpf",))
            ]
            if not active_drivers:
                for ident in identifiers:
                    detail_entry = {
                        "input": ident,
                        "type": id_type,
                        "operator": "",
                        "status": "erro",
                        "plan": "",
                        "message": "nenhum driver suporta este tipo",
                        "captured_at": datetime.utcnow().isoformat(),
                        "debug": {"reason": "unsupported_id_type"},
                    }
                    results.append(detail_entry)
                    detailed_entries.append(detail_entry)
                    error += 1
                    processed += 1
                    job_logger.error(
                        "identifier_unsupported",
                        identifier=ident,
                        id_type=id_type,
                    )
                    await update_job_progress()
                continue

            for drv in active_drivers:
                driver_name = getattr(drv, "operator", getattr(drv, "name", "unknown"))
                logger.info(
                    f"🚀 Iniciando driver {driver_name} com {len(identifiers)} CPFs"
                )
                print(
                    f"[DEBUG] Rodando driver: {driver_name} - {len(identifiers)} CPFs"
                )

            expected = len(active_drivers)
            for ident in identifiers:
                identifier_meta[ident] = {"expected": expected, "id_type": id_type}

            await driver_manager.run_batch(
                identifiers,
                id_type,
                cache=cache,
                db=db,
                progress_callback=handle_progress,
            )

            for ident in list(identifiers):
                if ident in identifier_meta:
                    await finalize_identifier(ident)

        if results_buffer:
            for ident in list(results_buffer.keys()):
                await finalize_identifier(ident)

            for entry in entries:
                detail_entry = {
                    "input": identifier,
                    "type": meta.get("id_type", forced_type),
                    "operator": entry.operator,
                    "status": entry.status,
                    "plan": entry.plan,
                    "message": entry.message,
                    "captured_at": entry.captured_at,
                    "debug": entry.debug,
                }
                results.append(detail_entry)
                detailed_entries.append(detail_entry)

            await update_job_progress()
            job_logger.info(
                "identifier_processed",
                identifier=identifier,
                id_type=meta.get("id_type", forced_type),
                drivers=len(entries),
                success=has_success,
            )

        async def handle_progress(
            identifier: str, driver: BaseDriver, result: DriverResult, from_cache: bool
        ) -> None:
            meta = identifier_meta.get(identifier)
            if not meta:
                return
            debug_info = dict(result.debug or {})
            if from_cache:
                debug_info["cache_hit"] = True
            result.debug = debug_info
            job_logger.info(
                "driver_result",
                identifier=identifier,
                id_type=meta.get("id_type", forced_type),
                driver=driver.name,
                status=result.status,
                plan=result.plan,
                message=result.message,
                cached=from_cache,
                debug=debug_info,
            )
            results_buffer[identifier].append(result)
            if len(results_buffer[identifier]) >= meta.get("expected", 0):
                await finalize_identifier(identifier)

        for id_type, identifiers in grouped.items():
            if not identifiers:
                continue
            active_drivers = [
                drv
                for drv in drivers
                if id_type in getattr(drv, "supported_id_types", ("cpf",))
            ]
            if not active_drivers:
                for ident in identifiers:
                    detail_entry = {
                        "input": ident,
                        "type": id_type,
                        "operator": "",
                        "status": "erro",
                        "plan": "",
                        "message": "nenhum driver suporta este tipo",
                        "captured_at": datetime.utcnow().isoformat(),
                        "debug": {"reason": "unsupported_id_type"},
                    }
                    results.append(detail_entry)
                    detailed_entries.append(detail_entry)
                    error += 1
                    processed += 1
                    job_logger.error(
                        "identifier_unsupported",
                        identifier=ident,
                        id_type=id_type,
                    )
                    await update_job_progress()
                continue

            for drv in active_drivers:
                driver_name = getattr(drv, "operator", getattr(drv, "name", "unknown"))
                logger.info(
                    f"🚀 Iniciando driver {driver_name} com {len(identifiers)} CPFs"
                )
                print(
                    f"[DEBUG] Rodando driver: {driver_name} - {len(identifiers)} CPFs"
                )

            expected = len(active_drivers)
            for ident in identifiers:
                identifier_meta[ident] = {"expected": expected, "id_type": id_type}

            await driver_manager.run_batch(
                identifiers,
                id_type,
                cache=cache,
                db=db,
                progress_callback=handle_progress,
            )

            for ident in list(identifiers):
                if ident in identifier_meta:
                    await finalize_identifier(ident)

        if results_buffer:
            for ident in list(results_buffer.keys()):
                await finalize_identifier(ident)

        out_df = pd.DataFrame(results)
        xlsx_path = os.path.join(EXPORT_DIR, f"{job_id}.xlsx")
        build_xlsx_from_results(out_df, xlsx_path)

        if results:
            docs = []
            for row in results:
                doc = dict(row)
                doc["job_id"] = job_id
                docs.append(doc)
            if docs:
                await db.job_results.insert_many(docs)

        job_finished_at = datetime.utcnow().isoformat()

        write_last_run_log(
            job_id,
            total,
            success,
            error,
            None,
            xlsx_path,
            details=detailed_entries,
            started_at=job_started_at,
            finished_at=job_finished_at,
            job_type=forced_type,
            job_log_path=job_logger.path,
        )

        job_logger.info(
            "job_completed",
            total=total,
            success=success,
            error=error,
            xlsx_path=xlsx_path,
        )
        logger.info(f"[LIVE] Status atual do job: {job_id} - completed")
        if results:
            docs = []
            for row in results:
                doc = dict(row)
                doc["job_id"] = job_id
                docs.append(doc)
            if docs:
                await db.job_results.insert_many(docs)

        job_finished_at = datetime.utcnow().isoformat()

        write_last_run_log(
            job_id,
            total,
            success,
            error,
            None,
            xlsx_path,
            details=detailed_entries,
            started_at=job_started_at,
            finished_at=job_finished_at,
            job_type=forced_type,
            job_log_path=job_logger.path,
        )

        job_logger.info(
            "job_completed",
            total=total,
            success=success,
            error=error,
            xlsx_path=xlsx_path,
        )
        logger.info(f"[LIVE] Status atual do job: {job_id} - completed")

        await db.jobs.update_one(
            {"_id": job_id},
            {
                "$set": {
                    "status": "completed",
                    "total": total,
                    "success": success,
                    "error": error,
                    "processed": processed,
                    "completed_at": datetime.utcnow().isoformat(),
                    "xlsx_path": xlsx_path,
                    "job_log_path": job_logger.path,
                }
            },
        )
        await db.jobs.update_one(
            {"_id": job_id},
            {
                "$set": {
                    "status": "completed",
                    "total": total,
                    "success": success,
                    "error": error,
                    "processed": processed,
                    "completed_at": datetime.utcnow().isoformat(),
                    "xlsx_path": xlsx_path,
                    "job_log_path": job_logger.path,
                }
            },
        )
    except Exception as e:
        job_logger.error("job_failed", error=str(e))
        logger.info(f"[LIVE] Status atual do job: {job_id} - failed")
        write_last_run_log(
            job_id,
            0,
            0,
            0,
            None,
            None,
            error_message=str(e),
            details=detailed_entries,
            started_at=job_started_at,
            finished_at=datetime.utcnow().isoformat(),
            job_type=forced_type,
            job_log_path=job_logger.path,
        )
        await db.jobs.update_one(
            {"_id": job_id},
            {
                "$set": {
                    "status": "failed",
                    "error_message": str(e),
                    "completed_at": datetime.utcnow().isoformat(),
                    "job_log_path": job_logger.path,
                }
            },
        )
        job_logger.error("job_failed", error=str(e))
        logger.info(f"[LIVE] Status atual do job: {job_id} - failed")
        write_last_run_log(
            job_id,
            0,
            0,
            0,
            None,
            None,
            error_message=str(e),
            details=detailed_entries,
            started_at=job_started_at,
            finished_at=datetime.utcnow().isoformat(),
            job_type=forced_type,
            job_log_path=job_logger.path,
        )
        await db.jobs.update_one(
            {"_id": job_id},
            {
                "$set": {
                    "status": "failed",
                    "error_message": str(e),
                    "completed_at": datetime.utcnow().isoformat(),
                    "job_log_path": job_logger.path,
                }
            },
        )


def build_xlsx_from_results(df: pd.DataFrame, out_path: str):
    header = ["CPF", "amil", "bradesco", "unimed", "unimed seguros"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Consulta CPF"
    ws.append(header)
    by_cpf = {}
    for _, row in df.iterrows():
        if row.get("type") != "cpf":
            continue
        cpf_digits = str(row.get("input", ""))
        cpf_fmt = format_cpf(cpf_digits)
        op = str(row.get("operator", "")).lower()
        plan = str(row.get("plan", ""))
        status = str(row.get("status", ""))
        val = plan if plan else status
        if not cpf_fmt:
            continue
        if cpf_fmt not in by_cpf:
            by_cpf[cpf_fmt] = {}
        by_cpf[cpf_fmt][op] = val
    for cpf_fmt, ops in by_cpf.items():
        row = [cpf_fmt, ops.get("amil", ""), ops.get("bradesco", ""), ops.get("unimed", ""), ops.get("seguros_unimed", "")]
        ws.append(row)
    for cell in ws["A"]:
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")
    wb.save(out_path)


def write_last_run_log(
    job_id: str,
    total: int,
    success: int,
    error: int,
    csv_path: Optional[str],
    xlsx_path: Optional[str],
    *,
    error_message: Optional[str] = None,
    details: Optional[List[Dict[str, Any]]] = None,
    started_at: Optional[str] = None,
    finished_at: Optional[str] = None,
    job_type: str = "auto",
    job_log_path: Optional[str] = None,
) -> None:
    lines = [
        f"job_id: {job_id}",
        f"job_type: {job_type}",
        f"started_at: {started_at or ''}",
        f"finished_at: {finished_at or ''}",
        f"job_type: {job_type}",
        f"started_at: {started_at or ''}",
        f"finished_at: {finished_at or ''}",
        f"total: {total}",
        f"success: {success}",
        f"error: {error}",
    ]
    if csv_path:
        lines.append(f"csv: {csv_path}")
    if xlsx_path:
        lines.append(f"xlsx: {xlsx_path}")
    if job_log_path:
        lines.append(f"job_log: {job_log_path}")
    if job_log_path:
        lines.append(f"job_log: {job_log_path}")
    if error_message:
        lines.append(f"error_message: {error_message}")

    if details:
        lines.append("--- details ---")
        for entry in details:
            inp = entry.get("input", "")
            ident_type = entry.get("type", "")
            operator = entry.get("operator", "")
            status = entry.get("status", "")
            plan = entry.get("plan", "")
            message = entry.get("message", "") or ""
            lines.append(f"- input: {inp} ({ident_type})")
            lines.append(f"  operator: {operator} | status: {status} | plan: {plan}")
            if message:
                lines.append(f"  message: {message}")

            debug = entry.get("debug") or {}
            if isinstance(debug, dict) and debug:
                reason = debug.get("reason")
                if reason:
                    lines.append(f"  reason: {reason}")
                captured = debug.get("captured_text")
                if captured:
                    lines.append(f"  captured_text: {captured[:300]}")
                status_selector = debug.get("status_selector")
                if status_selector:
                    lines.append(f"  status_selector: {status_selector}")
                plan_selector = debug.get("plan_selector")
                if plan_selector:
                    lines.append(f"  plan_selector: {plan_selector}")
                plan_text = debug.get("plan_text")
                if plan_text:
                    lines.append(f"  plan_text: {plan_text[:300]}")
                decided_status = debug.get("decided_status")
                if decided_status and decided_status != status:
                    lines.append(f"  decided_status: {decided_status}")
                error_detail = debug.get("error")
                if error_detail:
                    lines.append(f"  debug_error: {error_detail}")
                artifacts = debug.get("artifacts")
                if isinstance(artifacts, dict):
                    for key, value in artifacts.items():
                        lines.append(f"  artifact_{key}: {value}")
                steps = debug.get("steps")
                if isinstance(steps, list) and steps:
                    lines.append("  steps:")
                    for step in steps:
                        idx = step.get("index")
                        action = step.get("action")
                        selector = step.get("selector") or step.get("target") or ""
                        step_status = step.get("status")
                        lines.append(
                            f"    - #{idx} {action or ''} {selector} status={step_status}"
                        )
                        if step.get("error"):
                            lines.append(f"      error: {step['error']}")


    if details:
        lines.append("--- details ---")
        for entry in details:
            inp = entry.get("input", "")
            ident_type = entry.get("type", "")
            operator = entry.get("operator", "")
            status = entry.get("status", "")
            plan = entry.get("plan", "")
            message = entry.get("message", "") or ""
            lines.append(f"- input: {inp} ({ident_type})")
            lines.append(f"  operator: {operator} | status: {status} | plan: {plan}")
            if message:
                lines.append(f"  message: {message}")

            debug = entry.get("debug") or {}
            if isinstance(debug, dict) and debug:
                reason = debug.get("reason")
                if reason:
                    lines.append(f"  reason: {reason}")
                captured = debug.get("captured_text")
                if captured:
                    lines.append(f"  captured_text: {captured[:300]}")
                status_selector = debug.get("status_selector")
                if status_selector:
                    lines.append(f"  status_selector: {status_selector}")
                plan_selector = debug.get("plan_selector")
                if plan_selector:
                    lines.append(f"  plan_selector: {plan_selector}")
                plan_text = debug.get("plan_text")
                if plan_text:
                    lines.append(f"  plan_text: {plan_text[:300]}")
                decided_status = debug.get("decided_status")
                if decided_status and decided_status != status:
                    lines.append(f"  decided_status: {decided_status}")
                error_detail = debug.get("error")
                if error_detail:
                    lines.append(f"  debug_error: {error_detail}")
                artifacts = debug.get("artifacts")
                if isinstance(artifacts, dict):
                    for key, value in artifacts.items():
                        lines.append(f"  artifact_{key}: {value}")
                steps = debug.get("steps")
                if isinstance(steps, list) and steps:
                    lines.append("  steps:")
                    for step in steps:
                        idx = step.get("index")
                        action = step.get("action")
                        selector = step.get("selector") or step.get("target") or ""
                        step_status = step.get("status")
                        lines.append(
                            f"    - #{idx} {action or ''} {selector} status={step_status}"
                        )
                        if step.get("error"):
                            lines.append(f"      error: {step['error']}")

    os.makedirs(LOGS_DIR, exist_ok=True)
    with open(LAST_RUN_LOG, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
